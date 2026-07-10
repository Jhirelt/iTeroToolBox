"""
iTero Support Toolbox - Backend
All functions run as the current user (app must be launched as Administrator).
"""

import subprocess
import os
import sys
import json
import re
import time
import ctypes
import ctypes.wintypes
import winreg
import socket
import threading
from datetime import datetime
from pathlib import Path


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _run(cmd, shell=True, timeout=30):
    """Run a command, return (stdout, stderr, returncode)."""
    try:
        result = subprocess.run(
            cmd,
            shell=shell,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding="utf-8",
            errors="replace"
        )
        return result.stdout.strip(), result.stderr.strip(), result.returncode
    except subprocess.TimeoutExpired:
        return "", "Command timed out", 1
    except Exception as e:
        return "", str(e), 1


def _run_elevated(cmd, wait=True):
    """
    Run a command elevated (ShellExecute runas).
    For commands that need UAC even when already admin.
    """
    try:
        import win32api
        import win32con
        ret = win32api.ShellExecute(
            0, "runas", "cmd.exe",
            f'/c "{cmd}"', None, win32con.SW_HIDE if not wait else win32con.SW_SHOW
        )
        return str(ret), "", 0
    except Exception as e:
        return "", str(e), 1


def _ts():
    return datetime.now().strftime("%H:%M:%S")


def ok(data=None, msg=""):
    return json.dumps({"status": "ok", "data": data, "msg": msg, "ts": _ts()})


def err(msg):
    return json.dumps({"status": "error", "data": None, "msg": msg, "ts": _ts()})


# ─────────────────────────────────────────────
# HEALTH CHECK
# ─────────────────────────────────────────────

def health_check():
    """Run all health checks and return results dict."""
    results = {}

    # Ping gateway
    try:
        gw_out, _, _ = _run("ipconfig | findstr /i \"Default Gateway\"")
        gw = None
        for line in gw_out.splitlines():
            parts = line.split(":")
            if len(parts) == 2:
                ip = parts[1].strip()
                if ip and ip != "":
                    gw = ip
                    break
        if gw:
            ping_out, _, rc = _run(f"ping -n 2 -w 1000 {gw}")
            if rc == 0:
                match = re.search(r"Average = (\d+)ms", ping_out)
                ms = match.group(1) if match else "OK"
                results["ping"] = {"status": "ok", "val": f"{ms} ms", "detail": gw}
            else:
                results["ping"] = {"status": "error", "val": "No reply", "detail": gw}
        else:
            results["ping"] = {"status": "warn", "val": "No gateway", "detail": ""}
    except Exception as e:
        results["ping"] = {"status": "error", "val": "Error", "detail": str(e)}

    # DNS
    try:
        out, _, rc = _run("ping -n 1 -w 2000 google.com")
        if rc == 0:
            results["dns"] = {"status": "ok", "val": "OK", "detail": "google.com resolved"}
        else:
            results["dns"] = {"status": "error", "val": "Failed", "detail": "Cannot resolve google.com"}
    except Exception as e:
        results["dns"] = {"status": "error", "val": "Error", "detail": str(e)}

    # iTero service
    try:
        out, _, rc = _run("sc query ITeroTransferService")
        if "RUNNING" in out:
            results["svc"] = {"status": "ok", "val": "Running", "detail": "ITeroTransferService"}
        elif "STOPPED" in out:
            results["svc"] = {"status": "warn", "val": "Stopped", "detail": "ITeroTransferService"}
        else:
            results["svc"] = {"status": "warn", "val": "Not found", "detail": "Service missing"}
    except Exception as e:
        results["svc"] = {"status": "warn", "val": "Unknown", "detail": str(e)}

    # IPv6
    try:
        out, _, _ = _run("netsh int ipv6 show global")
        if "disabled" in out.lower():
            results["ipv6"] = {"status": "ok", "val": "Disabled", "detail": "IPv6 off"}
        else:
            results["ipv6"] = {"status": "warn", "val": "Enabled", "detail": "Consider disabling"}
    except Exception as e:
        results["ipv6"] = {"status": "warn", "val": "Unknown", "detail": str(e)}

    # Battery
    try:
        out, _, rc = _run(
            'powershell -NoProfile -Command '
            '"(Get-WmiObject -Class Win32_Battery).EstimatedChargeRemaining"'
        )
        if rc == 0 and out.strip().isdigit():
            pct = int(out.strip())
            status = "ok" if pct > 30 else ("warn" if pct > 15 else "error")
            results["bat"] = {"status": status, "val": f"{pct}%", "detail": "Battery charge"}
        else:
            results["bat"] = {"status": "warn", "val": "N/A", "detail": "No battery / AC only"}
    except Exception as e:
        results["bat"] = {"status": "warn", "val": "Error", "detail": str(e)}

    # Disk — use Python shutil for reliable C: free space
    try:
        import shutil as _shutil
        _total, _used, _free = _shutil.disk_usage("C:\\")
        _val = round(_free / (1024**3), 1)
        _status = "ok" if _val > 10 else ("warn" if _val > 5 else "error")
        results["disk"] = {"status": _status, "val": f"{_val} GB free", "detail": "C: drive"}
    except Exception as e:
        results["disk"] = {"status": "warn", "val": "Error", "detail": str(e)}

    # RAM
    try:
        out, _, _ = _run(
            'powershell -NoProfile -Command '
            '"$os=Get-WmiObject Win32_OperatingSystem; '
            '[Math]::Round(($os.TotalVisibleMemorySize-$os.FreePhysicalMemory)/$os.TotalVisibleMemorySize*100)"'
        )
        pct = out.strip()
        try:
            val = int(pct)
            status = "ok" if val < 80 else ("warn" if val < 90 else "error")
            results["ram"] = {"status": status, "val": f"{pct}% used", "detail": "RAM usage"}
        except Exception:
            results["ram"] = {"status": "warn", "val": pct or "?", "detail": "RAM usage"}
    except Exception as e:
        results["ram"] = {"status": "warn", "val": "Error", "detail": str(e)}

    # OS
    try:
        out, _, _ = _run(
            'powershell -NoProfile -Command '
            '"(Get-WmiObject Win32_OperatingSystem).Caption + \' \' + '
            '(Get-WmiObject Win32_OperatingSystem).BuildNumber"'
        )
        results["os2"] = {"status": "ok", "val": out.strip() or "Windows", "detail": "OS version"}
    except Exception as e:
        results["os2"] = {"status": "ok", "val": "Windows", "detail": str(e)}

    return json.dumps({"status": "ok", "data": results, "ts": _ts()})


# ─────────────────────────────────────────────
# NETWORKING
# ─────────────────────────────────────────────

def run_ipconfig():
    out, stderr, rc = _run("ipconfig /all", timeout=15)
    if rc == 0:
        # Chunk output — pywebview has a JS string size limit
        return ok(out[:8000] if len(out) > 8000 else out)
    return err(stderr)


def run_ping(target="8.8.8.8"):
    out, stderr, rc = _run(f"ping -n 4 {target}")
    return ok(out) if rc == 0 else err(out or stderr)


def get_wifi_details():
    out, stderr, rc = _run("netsh wlan show interfaces")
    return ok(out) if rc == 0 else err(stderr)


def get_network_details():
    """Return structured network info for Transfer Service panel."""
    out, _, _ = _run("ipconfig /all")
    data = {
        "ssid": "", "mac": "", "ipv4": "", "subnet": "",
        "gateway": "", "dhcp": "", "dns": "", "dhcp_enabled": ""
    }
    # Parse relevant fields
    lines = out.splitlines()
    in_wifi = False
    for line in lines:
        l = line.lower()
        if "wireless" in l or "wi-fi" in l or "wlan" in l:
            in_wifi = True
        if in_wifi:
            if "physical address" in l:
                data["mac"] = line.split(":")[-1].strip() if ":" in line else ""
            elif "ipv4 address" in l and not data["ipv4"]:
                data["ipv4"] = re.sub(r"\(.*\)", "", line.split(":")[-1]).strip()
            elif "subnet mask" in l:
                data["subnet"] = line.split(":")[-1].strip()
            elif "default gateway" in l:
                data["gateway"] = line.split(":")[-1].strip()
            elif "dhcp server" in l:
                data["dhcp"] = line.split(":")[-1].strip()
            elif "dhcp enabled" in l:
                data["dhcp_enabled"] = line.split(":")[-1].strip()
            elif "dns servers" in l:
                data["dns"] = line.split(":")[-1].strip()
    # SSID
    wifi_out, _, _ = _run("netsh wlan show interfaces")
    for line in wifi_out.splitlines():
        if "ssid" in line.lower() and "bssid" not in line.lower():
            data["ssid"] = line.split(":")[-1].strip()
            break
    return ok(data)


def disable_ipv6():
    out, stderr, rc = _run("netsh int ipv6 set state disabled")
    if rc == 0:
        return ok("IPv6 disabled successfully")
    return err(stderr or out)


def flush_dns():
    out, stderr, rc = _run("ipconfig /flushdns")
    return ok(out) if rc == 0 else err(stderr)


def generate_network_report():
    """Generate netsh network report and return path."""
    path = r"C:\Users\Public\network-report.html"
    out, stderr, rc = _run(f'netsh wlan show wlanreport output-file="{path}"')
    if os.path.exists(path):
        return ok({"path": path, "output": out})
    # fallback: netsh trace
    return ok({"path": "", "output": out or stderr})


# ─────────────────────────────────────────────
# BATTERY
# ─────────────────────────────────────────────

def generate_battery_report():
    path = r"C:\Users\iTero\Documents\battery-report.html"
    out, stderr, rc = _run(
        f'powercfg /batteryreport /duration 14 /output "{path}"'
    )
    if os.path.exists(path):
        return ok({"path": path, "output": out})
    # Try public path as fallback
    path2 = r"C:\Users\Public\battery-report.html"
    out2, _, _ = _run(f'powercfg /batteryreport /duration 14 /output "{path2}"')
    if os.path.exists(path2):
        return ok({"path": path2, "output": out2})
    return err(f"Report not found. {stderr}")


def read_battery_report(path):
    """Read battery report HTML and extract design/full capacity."""
    try:
        if not os.path.exists(path):
            return err("File not found")
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            html = f.read()
        design = re.search(r"DESIGN CAPACITY.*?(\d[\d,]+)\s*mWh", html, re.IGNORECASE)
        full = re.search(r"FULL CHARGE CAPACITY.*?(\d[\d,]+)\s*mWh", html, re.IGNORECASE)
        cycle = re.search(r"CYCLE COUNT.*?(\d+)", html, re.IGNORECASE)
        d_cap = design.group(1).replace(",", "") if design else None
        f_cap = full.group(1).replace(",", "") if full else None
        cycles = cycle.group(1) if cycle else None
        health = None
        if d_cap and f_cap:
            try:
                health = round(int(f_cap) / int(d_cap) * 100, 1)
            except Exception:
                pass
        return ok({
            "design_capacity": d_cap,
            "full_capacity": f_cap,
            "cycle_count": cycles,
            "health_pct": health,
            "path": path
        })
    except Exception as e:
        return err(str(e))


def get_battery_events():
    cmd = (
        'powershell -NoProfile -Command '
        '"Get-WinEvent -FilterHashtable @{LogName=\'System\'; Id=41,12,105} '
        '-MaxEvents 15 -ErrorAction SilentlyContinue | '
        'Select-Object TimeCreated,Id,Message | ConvertTo-Json"'
    )
    out, stderr, rc = _run(cmd, timeout=20)
    return ok(out) if out else err(stderr or "No events found")


# ─────────────────────────────────────────────
# SERVICES
# ─────────────────────────────────────────────

def check_service(name):
    out, stderr, rc = _run(f"sc query {name}")
    if "RUNNING" in out:
        status = "running"
    elif "STOPPED" in out:
        status = "stopped"
    elif "does not exist" in out.lower() or rc != 0:
        status = "missing"
    else:
        status = "unknown"
    return ok({"name": name, "status": status, "raw": out})


def set_service_auto_start(name):
    out1, _, rc1 = _run(f"sc config {name} start= auto")
    out2, _, rc2 = _run(f"sc start {name}")
    if rc1 == 0:
        return ok(f"Service {name} set to Automatic. Start result: {out2}")
    return err(f"Failed to configure service: {out1}")


# ─────────────────────────────────────────────
# REGISTRY
# ─────────────────────────────────────────────

def query_registry(key_path):
    """Query a registry key and return its values."""
    out, stderr, rc = _run(f'reg query "{key_path}"')
    return ok(out) if rc == 0 else err(stderr or out)


def set_registry_value(key_path, value_name, value_data, value_type="REG_BINARY"):
    out, stderr, rc = _run(
        f'reg add "{key_path}" /v "{value_name}" /t {value_type} /d "{value_data}" /f'
    )
    return ok(out) if rc == 0 else err(stderr or out)


def open_regedit(key_path=None):
    """Open regedit, optionally pre-navigated to a key."""
    if key_path:
        try:
            reg_key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Applets\Regedit"
            )
            winreg.SetValueEx(reg_key, "LastKey", 0, winreg.REG_SZ, key_path)
            winreg.CloseKey(reg_key)
        except Exception:
            pass
    subprocess.Popen(["regedit.exe"])
    return ok("Registry Editor opened")


# ─────────────────────────────────────────────
# KILL iTERO
# ─────────────────────────────────────────────

ITERO_PROCESSES = [
    "Cadent.iTero.Shell.iTeroScannerShell.exe",
    "Cadent.iTero.iTeroDesktopApplication.exe",
    "Cadent.iTero.UI.Application.exe",
    "RxManager.WinFormHostApp.exe",
    "RxManager.TrayIconApplication.exe",
    "Aligntech.RTH.MatServerInterface.exe",
    "CaseAdmin.exe",
    "Cadent.CaseAdmin.Application.exe",
    "loguploader.exe",
    "HwDiagnosticsApplicationUI.exe",
    "RunHwDiagnosticsApplication.exe",
    "Cadent.Utilities.KeyMonitor.exe",
    "Cadent.Infra.Legacy.minidump.exe",
    "Cadent.Shared.Util.Watchdog.exe",
    "EnableKbFilter.exe",
    "Align.Utilities.UserActivityDetection.exe",
    "Cadent.Shared.Util.DrawingToolWeb.exe",
    "Cadent.iTero.UI.SplashScreen.exe",
    "Align.iTero.Packaging.Util.Packager.exe",
    "RxManager.WpfHostApp.exe",
    "CefSharp.BrowserSubprocess.exe",
    "ITeroTransferService.exe",
    "InstallationApp.exe",
    "OutcomeSimulator.exe",
    "JarvisClient.exe",
    "DotNetRuntimeSetup.exe",
    "Angie.BG.Application.exe",
    "angie.tools.minidump_service.exe",
    "Scanapp.UI.Application.exe",
    "Cadent.Utilities.AngieBackgroundMock.exe",
    "Cadent.Utilities.AngieMock.exe",
    "Cadent.Utilities.SystemSnapshot.exe",
    "IAStorIcon.exe",
    "RAVCpl64.exe",
    "MageSystemMetricsService.exe",
]

def kill_all_itero():
    """Kill all iTero/Cadent processes and stop services."""
    killed = []
    failed = []
    for proc in ITERO_PROCESSES:
        out, _, rc = _run(f'taskkill /F /IM "{proc}"')
        if "SUCCESS" in out or rc == 0:
            killed.append(proc)
        else:
            failed.append(proc)
    # Stop services
    _run("sc stop winlogbeat")
    _run("sc stop filebeat")
    _run("sc stop MageSystemMetricsService")
    # PowerShell fallback for anything Cadent-pathed
    _run(
        'powershell -NoProfile -Command '
        '"Get-Process | ForEach-Object { try { if($_.Path -like \'*Cadent*\')'
        '{ Stop-Process -Id $_.Id -Force } } catch {} }"'
    )
    return ok({
        "killed": killed,
        "not_found": failed,
        "msg": f"Terminated {len(killed)} processes"
    })


def kill_process(process_name):
    out, _, rc = _run(f'taskkill /F /IM "{process_name}"')
    return ok(out) if "SUCCESS" in out or rc == 0 else err(out)


# ─────────────────────────────────────────────
# CLEAN INSTALL
# ─────────────────────────────────────────────

def rename_cadent_folders():
    cmd = (
        'powershell -NoProfile -ExecutionPolicy Bypass -Command '
        '"$paths=@(\'C:\\Program Files\\Cadent\','
        '\'C:\\Program Files (x86)\\Cadent\','
        '\'C:\\ProgramData\\Cadent\');'
        'foreach($p in $paths){'
        '  if(Test-Path $p){'
        '    $parent=Split-Path $p;'
        '    $new=Join-Path $parent \'Cadent_old\';'
        '    Rename-Item -Path $p -NewName \'Cadent_old\' -Force -ErrorAction SilentlyContinue;'
        '    Write-Host \'Renamed: \' $p'
        '  } else { Write-Host \'Not found: \' $p }'
        '}"'
    )
    out, stderr, rc = _run(cmd, timeout=60)
    return ok(out) if rc == 0 else err(stderr or out)


def clean_registry():
    keys = [
        (r"HKLM\SOFTWARE\Cadent", "iTeroPackageVersion"),
        (r"HKLM\SOFTWARE\Cadent", "TensorCUDA"),
        (r"HKLM\SOFTWARE\Cadent", "BiosUpdate"),
        (r"HKLM\SOFTWARE\Cadent", "DotNetVersion"),
        (r"HKLM\SOFTWARE\Cadent\ElementDemoDB", "Version"),
        (r"HKLM\SOFTWARE\Cadent\AngieDemoDB", "Version"),
    ]
    results = []
    for path, name in keys:
        out, _, rc = _run(f'reg delete "{path}" /v "{name}" /f')
        results.append(f"{'OK' if rc == 0 else 'Skip'}: {path}\\{name}")
    return ok("\n".join(results))


def remove_leftover_files():
    cmds = [
        r'del "C:\ProgramData\Cadent\iTero\RxData.sqlite" /f /q',
        r'del "C:\ProgramData\Cadent\Config\cadent_ini.dll" /f /q',
        r'del "C:\ProgramData\Cadent\Config\RxManagerConfig.ini.dll" /f /q',
    ]
    results = []
    for cmd in cmds:
        out, _, rc = _run(cmd)
        results.append(f"{'OK' if rc == 0 else 'Skip/NF'}: {cmd.split(chr(34))[1]}")
    return ok("\n".join(results))


# ─────────────────────────────────────────────
# TRANSFER SERVICE / FOLDER
# ─────────────────────────────────────────────

def verify_export_folder():
    path = r"C:\itero\export"
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
        return ok(f"Created {path}")
    return ok(f"Exists: {path}")


def share_export_folder():
    out, stderr, rc = _run(
        r'net share export="C:\itero\export" /GRANT:Everyone,FULL'
    )
    return ok(out) if rc == 0 else err(stderr or out)


def disable_password_sharing():
    out, stderr, rc = _run(
        r'reg add "HKLM\SYSTEM\CurrentControlSet\Control\Lsa" '
        r'/v "restrictanonymous" /t REG_DWORD /d 0 /f'
    )
    return ok(out) if rc == 0 else err(stderr)


# ─────────────────────────────────────────────
# SYSTEM TOOLS
# ─────────────────────────────────────────────

def open_device_manager():
    subprocess.Popen(["devmgmt.msc"], shell=True)
    return ok("Device Manager opened")


def open_event_viewer():
    subprocess.Popen(["eventvwr.msc"], shell=True)
    return ok("Event Viewer opened")


def open_services():
    subprocess.Popen(["services.msc"], shell=True)
    return ok("Services opened")


def open_log_collector(output_path=r"C:\Temp\Logs"):
    exe = r"C:\Program Files\Cadent\LogCollector\log_collector.exe"
    if not os.path.exists(exe):
        return err(f"Log collector not found: {exe}")
    subprocess.Popen([exe, output_path, "5", "0"])
    return ok(f"Log Collector launched → {output_path}")


def launch_slfdd_driver():
    inf = r"C:\Program Files\Cadent\SLFDD\Driver Installer\cyusb3.inf"
    if not os.path.exists(inf):
        return err(f"Driver not found: {inf}")
    out, stderr, rc = _run(f'pnputil /add-driver "{inf}" /install')
    return ok(out) if rc == 0 else err(stderr or out)


def launch_slfdd_utility():
    exe = r"C:\Program Files\Cadent\SLFDD\Utility\SlfddCSharp.exe"
    if not os.path.exists(exe):
        return err(f"SLFDD utility not found: {exe}")
    subprocess.Popen([exe])
    return ok("SLFDD utility launched")


def open_url_in_browser(url):
    import webbrowser
    webbrowser.open(url)
    return ok(f"Opened: {url}")


def open_html_file(path):
    if not os.path.exists(path):
        return err(f"File not found: {path}")
    import webbrowser
    webbrowser.open(f"file:///{path}")
    return ok(f"Opened: {path}")


def get_filter_events(event_ids="41,12,105", max_events=15):
    cmd = (
        f'powershell -NoProfile -Command '
        f'"Get-WinEvent -FilterHashtable @{{LogName=\'System\'; '
        f'Id={event_ids}}} -MaxEvents {max_events} -ErrorAction SilentlyContinue | '
        f'Select-Object TimeCreated,Id,LevelDisplayName,Message | '
        f'Format-List"'
    )
    out, stderr, rc = _run(cmd, timeout=20)
    return ok(out) if out else err(stderr or "No matching events found")


def run_arbitrary_cmd(cmd):
    """Execute any command and return output. App must run as admin."""
    out, stderr, rc = _run(cmd, timeout=60)
    combined = out + ("\n" + stderr if stderr else "")
    return json.dumps({
        "status": "ok" if rc == 0 else "error",
        "data": combined,
        "rc": rc,
        "ts": _ts()
    })


def open_woa_collage(agent_email=""):
    """Save WOA reference collage to known path and open File Explorer there."""
    import shutil, sys

    if getattr(sys, 'frozen', False):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).parent

    src = base / "wand_reference_WOA.png"
    dest_dir = Path(r"C:\iTeroToolbox\Reference\WOA")
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / "wand_reference_WOA.png"

    if not src.exists():
        alt = Path(__file__).parent / "wand_reference_WOA.png"
        if alt.exists():
            src = alt

    if src.exists():
        shutil.copy2(str(src), str(dest))

    if not dest.exists():
        return err(f"Collage not found. Place wand_reference_WOA.png in {dest_dir}")

    if agent_email:
        try:
            from PIL import Image, ImageDraw, ImageFont
            img = Image.open(str(dest))
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("C:\\Windows\\Fonts\\arial.ttf", 18)
            except Exception:
                font = ImageFont.load_default()
            text = f"Sent by: {agent_email}"
            draw.rectangle([8, img.height-34, 520, img.height-4], fill=(20,22,30))
            draw.text((12, img.height-30), text, fill=(100,200,255), font=font)
            img.save(str(dest))
        except Exception:
            pass

    subprocess.Popen(f'explorer /select,"{dest}"', shell=True)
    return ok({"path": str(dest), "msg": f"File Explorer opened → {dest}"})


def generate_link_unlink_excel(data):
    """Generate compact Link/Unlink Excel — MAT, Unlink, Link only."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        mat    = data.get('mat', '')
        unlink = data.get('unlink', '')
        link   = data.get('link', '')
        agent  = data.get('agent', '') or 'Agent'
        date_s = datetime.now().strftime('%Y-%m-%d')
        ts     = datetime.now().strftime('%Y%m%d_%H%M%S')

        dest_dir = Path(r'C:\iTeroToolbox\LinkUnlink')
        dest_dir.mkdir(parents=True, exist_ok=True)
        fname = f'LinkUnlink_{ts}.xlsx'
        dest  = dest_dir / fname

        wb = Workbook()
        ws = wb.active
        ws.title = 'Link-Unlink'

        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        left   = Alignment(horizontal='left',   vertical='center')

        # Title
        ws.merge_cells('A1:C1')
        ws['A1'].value     = f'iTero Link/Unlink Request — {agent} — {date_s}'
        ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='1F3A6E')
        ws['A1'].alignment = center

        # Headers
        headers = ['MAT Person', 'SFDC Contact ID', 'Action']
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill('solid', start_color='1F3A6E')
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = border

        # Unlink row
        unlink_fill = PatternFill('solid', start_color='FFDCDC')
        for col, val in enumerate([mat, unlink, 'Unlink'], 1):
            c = ws.cell(row=4, column=col, value=val)
            c.fill = unlink_fill; c.border = border; c.alignment = left
            c.font = Font(name='Arial', bold=(col==3), color='C0392B' if col==3 else '000000', size=11)

        # Link row
        link_fill = PatternFill('solid', start_color='DCFFE4')
        for col, val in enumerate([mat, link, 'Link'], 1):
            c = ws.cell(row=5, column=col, value=val)
            c.fill = link_fill; c.border = border; c.alignment = left
            c.font = Font(name='Arial', bold=(col==3), color='1E7E34' if col==3 else '000000', size=11)

        # Column widths
        for i, w in enumerate([20, 24, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[3].height = 20

        wb.save(str(dest))
        subprocess.Popen(f'explorer /select,"{dest}"', shell=True)
        return ok({'path': str(dest), 'filename': fname})

    except Exception as e:
        return err(f'Failed to generate Excel: {e}')



def open_woa_collage_folder(path):
    """Open File Explorer selecting the given file path."""
    try:
        if os.path.exists(path):
            subprocess.Popen(f'explorer /select,"{path}"', shell=True)
        else:
            folder = str(Path(path).parent)
            subprocess.Popen(f'explorer "{folder}"', shell=True)
        return ok(f'Explorer opened: {path}')
    except Exception as e:
        return err(str(e))


def open_collage_by_b64(collage_type, b64_data):
    """Save a base64 collage image to disk and open in Explorer."""
    import base64, shutil
    try:
        # Strip data URI prefix if present
        if ',' in b64_data:
            b64_data = b64_data.split(',', 1)[1]
        img_bytes = base64.b64decode(b64_data)
        dest_dir = Path(r'C:\iTeroToolbox\LinkUnlink')
        dest_dir.mkdir(parents=True, exist_ok=True)
        names = {
            'yucblx': 'wand_reference_YUC_BLX.png',
            'lumina': 'wand_reference_LUMINA.png',
            'woa':    'wand_reference_WOA.png',
        }
        fname = names.get(collage_type, f'wand_reference_{collage_type}.png')
        dest = dest_dir / fname
        with open(str(dest), 'wb') as f:
            f.write(img_bytes)
        subprocess.Popen(f'explorer /select,"{dest}"', shell=True)
        return ok({'path': str(dest), 'msg': f'Explorer opened → {dest}'})
    except Exception as e:
        return err(str(e))


def open_collage_file(collage_type):
    """Open a pre-saved collage PNG from Reference\\subfolder in Explorer."""
    config = {
        'yucblx': (r'C:\iTeroToolbox\Reference\YUC_BLX', 'wand_reference_YUC_BLX.png'),
        'lumina':  (r'C:\iTeroToolbox\Reference\Lumina',  'wand_reference_LUMINA.png'),
        'woa':     (r'C:\iTeroToolbox\Reference\WOA',     'wand_reference_WOA.png'),
    }
    if collage_type not in config:
        return err(f'Unknown collage type: {collage_type}')
    folder, fname = config[collage_type]
    dest = Path(folder) / fname
    # If not in Reference subfolder, fall back to app dir
    if not dest.exists():
        alt = Path(__file__).parent / fname
        if alt.exists():
            # Copy to correct subfolder
            import shutil
            Path(folder).mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(alt), str(dest))
        else:
            return err(f'Collage not found: {fname}. Place PNG in {folder}\\')
    subprocess.Popen(f'explorer /select,"{dest}"', shell=True)
    return ok({'path': str(dest), 'msg': f'Explorer opened -> {dest}'})


def get_catalog_parts(model=""):
    """Return list of parts from C:\\iTeroToolbox\\Catalog\\ as base64-encoded images."""
    import base64
    catalog_dir = Path(r"C:\iTeroToolbox\Catalog")
    if not catalog_dir.exists():
        return err(f"Catalog folder not found: {catalog_dir}")

    image_exts = {".jpg", ".jpeg", ".png", ".webp"}
    parts = []
    for f in sorted(catalog_dir.iterdir()):
        if f.suffix.lower() not in image_exts:
            continue
        stem = f.stem  # e.g. "Element Wand- 108540"
        # Parse "Part Name- CODE" or "Part Name - CODE" or just use full stem
        m = re.match(r'^(.+?)\s*[-–]\s*(\d+)\s*$', stem)
        if m:
            part_name = m.group(1).strip()
            part_code = m.group(2).strip()
        else:
            part_name = stem
            part_code = ""
        try:
            b64 = base64.b64encode(f.read_bytes()).decode("utf-8")
            parts.append({
                "part_name": part_name,
                "part_code": part_code,
                "filename": f.name,
                "path": str(f),
                "b64": b64,
                "error": False
            })
        except Exception as e:
            parts.append({
                "part_name": part_name,
                "part_code": part_code,
                "filename": f.name,
                "path": str(f),
                "b64": "",
                "error": True
            })

    return ok(parts)


def fetch_sf_case_title(url=""):
    """Extract case ID from SF URL — title requires authenticated browser session."""
    if not url:
        return err("No URL provided")
    m = re.search(r'/([0-9]{8,18})', url)
    case_id = m.group(1) if m else ""
    # Salesforce requires an authenticated browser session — cannot fetch title server-side.
    return err(f"Salesforce requires login — fill title manually.{(' Case ID: ' + case_id) if case_id else ''}")


# ─────────────────────────────────────────────
# USER PROFILE
# ─────────────────────────────────────────────

_PROFILE_PATH = os.path.join(os.path.expandvars('%APPDATA%'), 'iTeroToolBox', 'profile.json')

def get_user_profile():
    try:
        if os.path.exists(_PROFILE_PATH):
            with open(_PROFILE_PATH, 'r') as f:
                return ok(json.load(f))
        return ok(None)
    except Exception as e:
        return err(str(e))

def save_user_profile(name="", role=""):
    try:
        os.makedirs(os.path.dirname(_PROFILE_PATH), exist_ok=True)
        data = {"name": name, "role": role}
        with open(_PROFILE_PATH, 'w') as f:
            json.dump(data, f)
        return ok(data)
    except Exception as e:
        return err(str(e))

def clear_user_profile():
    try:
        if os.path.exists(_PROFILE_PATH):
            os.remove(_PROFILE_PATH)
        return ok(True)
    except Exception as e:
        return err(str(e))


# ─────────────────────────────────────────────
# KB DATA (Subject Constructor — KB suggestions)
# ─────────────────────────────────────────────

_KB_DATA_CACHE = None

def get_kb_data():
    """Load data/kb_data.json once and cache it. Matching itself happens in JS."""
    global _KB_DATA_CACHE
    if _KB_DATA_CACHE is not None:
        return ok(_KB_DATA_CACHE)
    try:
        if getattr(sys, 'frozen', False):
            base = Path(sys._MEIPASS)
        else:
            base = Path(__file__).parent
        kb_path = base / "data" / "kb_data.json"
        with open(kb_path, 'r', encoding='utf-8') as f:
            _KB_DATA_CACHE = json.load(f)
        return ok(_KB_DATA_CACHE)
    except Exception as e:
        return err(str(e))


# ─────────────────────────────────────────────
# HOTKEYS  (global keyboard shortcuts via Win32 RegisterHotKey)
# ─────────────────────────────────────────────

_HK_PATH = os.path.join(os.path.expandvars('%APPDATA%'), 'iTeroToolBox', 'hotkeys.json')

_VK_MAP = {
    'F1':0x70,'F2':0x71,'F3':0x72,'F4':0x73,'F5':0x74,'F6':0x75,
    'F7':0x76,'F8':0x77,'F9':0x78,'F10':0x79,'F11':0x7A,'F12':0x7B,
    'F13':0x7C,'F14':0x7D,'F15':0x7E,'F16':0x7F,'F17':0x80,'F18':0x81,
    'F19':0x82,'F20':0x83,'F21':0x84,'F22':0x85,'F23':0x86,'F24':0x87,
    'A':0x41,'B':0x42,'C':0x43,'D':0x44,'E':0x45,'F':0x46,'G':0x47,
    'H':0x48,'I':0x49,'J':0x4A,'K':0x4B,'L':0x4C,'M':0x4D,'N':0x4E,
    'O':0x4F,'P':0x50,'Q':0x51,'R':0x52,'S':0x53,'T':0x54,'U':0x55,
    'V':0x56,'W':0x57,'X':0x58,'Y':0x59,'Z':0x5A,
    '0':0x30,'1':0x31,'2':0x32,'3':0x33,'4':0x34,
    '5':0x35,'6':0x36,'7':0x37,'8':0x38,'9':0x39,
}
_MOD_MAP      = {'ctrl': 0x0002, 'alt': 0x0001, 'shift': 0x0004}
_MOD_NOREPEAT = 0x4000
_WM_HOTKEY    = 0x0312
_INPUT_KBD    = 1
_KF_UNICODE   = 0x0004
_KF_KEYUP     = 0x0002
_VK_RETURN    = 0x0D

class _KEYBDINPUT(ctypes.Structure):
    _fields_ = [
        ('wVk',         ctypes.c_ushort),
        ('wScan',       ctypes.c_ushort),
        ('dwFlags',     ctypes.c_uint),
        ('time',        ctypes.c_uint),
        ('dwExtraInfo', ctypes.c_size_t),
    ]

class _INPUT_UNION(ctypes.Union):
    _fields_ = [('ki', _KEYBDINPUT), ('_pad', ctypes.c_byte * 32)]

class _INPUT(ctypes.Structure):
    _fields_ = [('type', ctypes.c_uint), ('_', _INPUT_UNION)]

class _MSG(ctypes.Structure):
    _fields_ = [
        ('hwnd',    ctypes.c_size_t),
        ('message', ctypes.c_uint),
        ('wParam',  ctypes.c_size_t),
        ('lParam',  ctypes.c_ssize_t),
        ('time',    ctypes.c_uint),
        ('ptX',     ctypes.c_long),
        ('ptY',     ctypes.c_long),
    ]

def _type_text(text, multiline=True):
    """Inject Unicode text via SendInput. Works in TeamViewer remote sessions."""
    user32 = ctypes.windll.user32
    if not multiline:
        text = text.replace('\r', '').replace('\n', ' ')
    inputs = []
    for ch in text:
        if ch == '\r':
            continue
        if ch == '\n':
            for flags in (0, _KF_KEYUP):
                ki = _KEYBDINPUT(wVk=_VK_RETURN, wScan=0, dwFlags=flags, time=0, dwExtraInfo=0)
                inputs.append(_INPUT(type=_INPUT_KBD, _=_INPUT_UNION(ki=ki)))
        else:
            code = ord(ch)
            if code > 0xFFFF:
                continue
            for flags in (_KF_UNICODE, _KF_UNICODE | _KF_KEYUP):
                ki = _KEYBDINPUT(wVk=0, wScan=code, dwFlags=flags, time=0, dwExtraInfo=0)
                inputs.append(_INPUT(type=_INPUT_KBD, _=_INPUT_UNION(ki=ki)))
    if not inputs:
        return 0
    arr = (_INPUT * len(inputs))(*inputs)
    sent = user32.SendInput(len(inputs), arr, ctypes.sizeof(_INPUT))
    return sent

_hk_thread_ref = None
_hk_stop_evt   = threading.Event()
_hk_status     = {'registered': 0, 'failed': [], 'fired': 0, 'thread_alive': False}

def _hk_thread_func(hotkeys):
    global _hk_status
    user32   = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32

    # Force creation of this thread's Win32 message queue BEFORE RegisterHotKey.
    # Without this, RegisterHotKey may succeed but WM_HOTKEY is never delivered.
    _init_msg = _MSG()
    user32.PeekMessageW(ctypes.byref(_init_msg), None, 0, 0, 0)  # PM_NOREMOVE

    tid = kernel32.GetCurrentThreadId()
    print(f'[HK] Thread started (tid={tid})', flush=True)

    registered = []
    hk_map   = {}
    failed   = []
    reg_id   = 1

    for hk in hotkeys:
        if not hk.get('enabled', True):
            continue
        vk = _VK_MAP.get((hk.get('key') or '').upper())
        if not vk:
            continue
        mods = _MOD_NOREPEAT
        for m in (hk.get('mods') or []):
            mods |= _MOD_MAP.get(m, 0)
        label = hk.get('label') or hk.get('key', '?')
        # Retry up to 5 times — Win32 may still hold the handle briefly after
        # the previous thread's UnregisterHotKey if the thread restart is fast.
        ok_reg = False
        for attempt in range(5):
            ok_reg = user32.RegisterHotKey(None, reg_id, mods, vk)
            if ok_reg:
                break
            err_code = kernel32.GetLastError()
            if err_code == 1409 and attempt < 4:  # ERROR_HOTKEY_ALREADY_REGISTERED
                print(f'[HK] retry {attempt+1} for {label} (err=1409)…', flush=True)
                time.sleep(0.1)
            else:
                break
        if ok_reg:
            registered.append(reg_id)
            hk_map[reg_id] = hk
            reg_id += 1
            print(f'[HK] OK registered: {label} (win_id={reg_id-1})', flush=True)
        else:
            err_code = kernel32.GetLastError()
            entry = f'{label} (err={err_code})'
            failed.append(entry)
            print(f'[HK] FAILED register: {entry}', flush=True)

    _hk_status = {'registered': len(registered), 'failed': failed, 'fired': 0, 'thread_alive': True}
    print(f'[HK] Thread alive — {len(registered)} registered, {len(failed)} failed', flush=True)

    msg = _MSG()
    while not _hk_stop_evt.is_set():
        if user32.PeekMessageW(ctypes.byref(msg), None, _WM_HOTKEY, _WM_HOTKEY, 1):
            if msg.message == _WM_HOTKEY:
                hk = hk_map.get(msg.wParam)
                print(f'[HK] WM_HOTKEY fired: wParam={msg.wParam} → {hk.get("label") if hk else "unknown"}', flush=True)
                if hk:
                    _hk_status['fired'] += 1
                    time.sleep(0.08)
                    if hk.get('source') == 'cmd' and hk.get('cmd'):
                        cmd = hk['cmd']
                        print(f'[HK] running cmd: {cmd[:80]}', flush=True)
                        threading.Thread(
                            target=subprocess.run,
                            args=[cmd],
                            kwargs={'shell': True, 'capture_output': True},
                            daemon=True
                        ).start()
                    else:
                        sent = _type_text(hk.get('text') or '', multiline=hk.get('multiline', True))
                        print(f'[HK] SendInput sent {sent} events', flush=True)
        else:
            time.sleep(0.02)

    _hk_status['thread_alive'] = False
    for hk_id in registered:
        result = user32.UnregisterHotKey(None, hk_id)
        if not result:
            ue = kernel32.GetLastError()
            print(f'[HK] UnregisterHotKey({hk_id}) FAILED err={ue}', flush=True)
        else:
            print(f'[HK] UnregisterHotKey({hk_id}) OK', flush=True)
    print('[HK] Thread stopped', flush=True)

def start_hotkey_listener():
    global _hk_thread_ref, _hk_stop_evt
    stop_hotkey_listener()
    try:
        hotkeys = json.loads(open(_HK_PATH, encoding='utf-8').read()) if os.path.exists(_HK_PATH) else []
    except Exception:
        hotkeys = []
    if not any(h.get('enabled', True) and _VK_MAP.get((h.get('key') or '').upper()) for h in hotkeys):
        return
    _hk_stop_evt = threading.Event()
    _hk_thread_ref = threading.Thread(target=_hk_thread_func, args=(hotkeys,), daemon=True)
    _hk_thread_ref.start()

def stop_hotkey_listener():
    global _hk_thread_ref
    if _hk_thread_ref and _hk_thread_ref.is_alive():
        _hk_stop_evt.set()
        _hk_thread_ref.join(timeout=2.0)
    _hk_thread_ref = None
    # Give Win32 time to release hotkey handles before next RegisterHotKey call.
    time.sleep(0.15)

def get_hotkeys():
    try:
        if os.path.exists(_HK_PATH):
            with open(_HK_PATH, 'r', encoding='utf-8') as f:
                return ok(json.load(f))
        return ok([])
    except Exception as e:
        return err(str(e))

def save_hotkeys(data):
    try:
        hotkeys = json.loads(data) if isinstance(data, str) else data
        os.makedirs(os.path.dirname(_HK_PATH), exist_ok=True)
        with open(_HK_PATH, 'w', encoding='utf-8') as f:
            json.dump(hotkeys, f, ensure_ascii=False, indent=2)
        start_hotkey_listener()
        active = sum(1 for h in hotkeys if h.get('enabled', True))
        return ok(active)
    except Exception as e:
        return err(str(e))

def get_hotkey_status():
    running = _hk_thread_ref is not None and _hk_thread_ref.is_alive()
    return ok({
        'thread_running': running,
        'registered':     _hk_status.get('registered', 0),
        'failed':         _hk_status.get('failed', []),
        'fired':          _hk_status.get('fired', 0),
        'input_struct_size': ctypes.sizeof(_INPUT),
    })

def test_type_text(delay_sec=2):
    """Type a test string after a delay so the user can switch focus."""
    def _do():
        time.sleep(float(delay_sec))
        sent = _type_text('iTero_HK_TEST_OK')
        print(f'[HK] test_type_text: SendInput returned {sent}', flush=True)
    threading.Thread(target=_do, daemon=True).start()
    return ok(f'Will type test string in {delay_sec}s — switch focus to a text field now')


# ─────────────────────────────────────────────
# MAT (myaligntech.com) SCRAPER
# ─────────────────────────────────────────────

import base64

_MAT_CFG_PATH = os.path.join(os.environ.get("APPDATA", "C:\\Users"), "iTeroToolBox", "mat_config.json")
_mat_driver = None  # kept alive while browser is open in visible mode

_SCANNER_TAB_ID  = "ui-id-5"
_SCANNER_SN_ID   = "ctl00_body_tabsContainer_scannerTab_txtScanner"
_SCANNER_OK_ID   = "ctl00_body_tabsContainer_scannerTab_btnScanner"
_WAND_SN_ID      = "ctl00_body_txtEmbeddedHeadSerialIdentifier"
_SCANNER_SN_CONFIRMED_ID = "ctl00_body_txtSerialIdentifier"
_REG_ID          = "ctl00_body_txtRegToReadOnly"
_STATUS_ID       = "ctl00_body_ddlStatus"
_MODEL_PAGE_ID   = "ctl00_body_txtScannerModel"
_NOTES_ID        = "ctl00_body_txtNotes"
_SW_ROW_CSS      = "tr[class*='report-row']"


def mat_save_credentials(email, password_plain, show_browser=False, enabled=True):
    try:
        cfg = {
            "email":        email,
            "password_b64": base64.b64encode(password_plain.encode("utf-8")).decode(),
            "show_browser": bool(show_browser),
            "enabled":      bool(enabled),
        }
        os.makedirs(os.path.dirname(_MAT_CFG_PATH), exist_ok=True)
        with open(_MAT_CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f)
        return ok("Credentials saved.")
    except Exception as e:
        return err(str(e))


def mat_get_credentials():
    try:
        if not os.path.exists(_MAT_CFG_PATH):
            return ok({"email": "", "password": "", "show_browser": False, "enabled": False})
        with open(_MAT_CFG_PATH, encoding="utf-8") as f:
            cfg = json.load(f)
        password = ""
        if cfg.get("password_b64"):
            try:
                password = base64.b64decode(cfg["password_b64"]).decode("utf-8")
            except Exception:
                password = ""
        return ok({
            "email":        cfg.get("email", ""),
            "password":     password,
            "show_browser": cfg.get("show_browser", False),
            "enabled":      cfg.get("enabled", False),
        })
    except Exception as e:
        return err(str(e))


def mat_set_enabled(enabled):
    """Toggle MAT auto-scrape on/off without touching credentials."""
    try:
        cfg = {}
        if os.path.exists(_MAT_CFG_PATH):
            with open(_MAT_CFG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
        cfg["enabled"] = bool(enabled)
        os.makedirs(os.path.dirname(_MAT_CFG_PATH), exist_ok=True)
        with open(_MAT_CFG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f)
        return ok(bool(enabled))
    except Exception as e:
        return err(str(e))


def mat_scrape(serial):
    global _mat_driver
    try:
        if not os.path.exists(_MAT_CFG_PATH):
            return err("No MAT credentials. Configure in Settings → MAT.")
        with open(_MAT_CFG_PATH, encoding="utf-8") as f:
            cfg = json.load(f)
        email    = cfg.get("email", "")
        password = base64.b64decode(cfg.get("password_b64", "")).decode("utf-8")
        show     = cfg.get("show_browser", False)
        if not email or not password:
            return err("MAT credentials incomplete. Configure in Settings → MAT.")
    except Exception as e:
        return err(f"Config read error: {e}")

    # Close any leftover driver from previous run
    if _mat_driver is not None:
        try:
            _mat_driver.quit()
        except Exception:
            pass
        _mat_driver = None

    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.edge.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
    except ImportError:
        return err("selenium not installed. Run: pip install selenium")

    driver = None
    try:
        from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException

        opts = Options()
        if not show:
            opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_experimental_option("excludeSwitches", ["enable-logging"])

        driver = webdriver.Edge(options=opts)
        # Ignore stale refs — elements go stale during ASP.NET postbacks
        wait = WebDriverWait(driver, 25,
                             ignored_exceptions=[StaleElementReferenceException,
                                                 NoSuchElementException])

        # ── Login ──────────────────────────────────────────────
        LOGIN_URL = "https://myaligntech.com/login.aspx?auth=false&ReturnUrl=%2f"
        driver.get(LOGIN_URL)
        wait.until(EC.presence_of_element_located((By.ID, "LoginControl_UserName")))
        driver.find_element(By.ID, "LoginControl_UserName").clear()
        driver.find_element(By.ID, "LoginControl_UserName").send_keys(email)
        driver.find_element(By.ID, "LoginControl_Password").clear()
        driver.find_element(By.ID, "LoginControl_Password").send_keys(password)
        # Use JS click on login button to avoid any overlay issues
        driver.execute_script(
            "document.getElementById('LoginControl_LoginButton').click();"
        )

        wait.until(lambda d: "login.aspx" not in d.current_url)
        time.sleep(0.8)

        if "login.aspx" in driver.current_url.lower():
            if show: _mat_driver = driver
            else: driver.quit()
            return err("MAT login failed — check credentials in Settings → MAT.")

        # ── Click Scanner tab via JS (jQuery UI anchor with tabindex=-1) ───
        wait.until(EC.presence_of_element_located((By.ID, _SCANNER_TAB_ID)))
        driver.execute_script(
            "document.getElementById(arguments[0]).click();", _SCANNER_TAB_ID
        )
        # Wait for the scanner SN input to become visible inside the tab panel
        wait.until(EC.visibility_of_element_located((By.ID, _SCANNER_SN_ID)))
        time.sleep(0.4)

        # ── Fill scanner SN and click OK via JS ────────────────
        driver.execute_script(
            "var f=document.getElementById(arguments[0]);f.value=arguments[1];",
            _SCANNER_SN_ID, serial
        )
        driver.execute_script(
            "document.getElementById(arguments[0]).click();", _SCANNER_OK_ID
        )

        # ── Wait for postback / UpdatePanel to finish ──────────
        # Strategy: poll until wand SN field has a non-empty value.
        # StaleElementReferenceException is ignored by our wait instance,
        # so we safely re-find the element on every poll tick.
        time.sleep(1.5)

        def wand_ready(d):
            el = d.find_element(By.ID, _WAND_SN_ID)
            return bool(el.get_attribute("value").strip())

        wand_sn = ""
        try:
            wait.until(wand_ready)
            wand_sn = driver.find_element(By.ID, _WAND_SN_ID).get_attribute("value").strip()
        except Exception as wait_err:
            # Postback may have completed but field is genuinely empty (SN not found).
            # Try reading anyway before giving up.
            try:
                wand_sn = driver.find_element(By.ID, _WAND_SN_ID).get_attribute("value").strip()
            except Exception:
                wand_sn = ""

        # ── Confirmed scanner S/N (as MAT itself reports it back, not just
        #    an echo of what we searched with — used for the S/N cross-check
        #    against the ticket and the Salesforce Subject) ────────────────
        scanner_sn = ""
        try:
            scanner_sn = driver.find_element(By.ID, _SCANNER_SN_CONFIRMED_ID).get_attribute("value").strip()
        except Exception:
            pass

        reg_to = ""
        try:
            reg_to = driver.find_element(By.ID, _REG_ID).get_attribute("value").strip()
        except Exception:
            pass

        # ── Status (Active / Inactive) ─────────────────────────
        status = ""
        try:
            from selenium.webdriver.support.ui import Select as SeleniumSelect
            sel_el  = driver.find_element(By.ID, _STATUS_ID)
            sel_val = SeleniumSelect(sel_el).first_selected_option.get_attribute("value")
            status  = "Active" if sel_val == "0" else "Inactive"
        except Exception:
            pass

        # ── Scanner model ──────────────────────────────────────
        scanner_model = ""
        try:
            scanner_model = driver.find_element(By.ID, _MODEL_PAGE_ID).get_attribute("value").strip()
        except Exception:
            pass

        # ── Notes ──────────────────────────────────────────────
        notes = ""
        try:
            notes = driver.find_element(By.ID, _NOTES_ID).get_attribute("value").strip()
        except Exception:
            pass

        # ── SW Package (first report-row: pkg=td[2], expiry=td[4]) ──
        sw_package = ""
        sw_expiry  = ""
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, _SW_ROW_CSS)
            for row in rows:
                tds = row.find_elements(By.TAG_NAME, "td")
                if len(tds) < 5:
                    continue
                # td[1] holds the row code — individual SW options are
                # "SOP-xxxxx" (empty dates), the actual package/bundle
                # row is "BUN-xxxxx" and is the only one with real dates.
                code = (tds[1].get_attribute("textContent") or "").strip()
                if not code.upper().startswith("BUN"):
                    continue
                # Use textContent, not .text — the row's container may be
                # hidden (display:none) until the user clicks "editar" in
                # the real UI, and .text returns "" for hidden elements.
                sw_package = (tds[2].get_attribute("textContent") or "").strip()
                sw_expiry  = (tds[4].get_attribute("textContent") or "").strip()
                break
        except Exception:
            pass

        # ── Company info (click "Reg. To" → Business Partner page) ──
        company_id      = ""
        company_name    = ""
        company_address = ""
        try:
            reg_link = driver.find_element(By.CSS_SELECTOR, "a[href*='BusinessPartnerPage.aspx']")
            driver.execute_script("arguments[0].click();", reg_link)
            wait.until(EC.presence_of_element_located((By.ID, "ctl00_body_EntityNameLabel")))
            name_text = driver.find_element(By.ID, "ctl00_body_EntityNameLabel").text.strip()
            company_address = driver.find_element(By.ID, "ctl00_body_EntityDescriptionLabel").text.strip()
            # name_text looks like "[Company ID: 112940] Dr. Sepideh Badri"
            m = re.search(r"\[Company ID:\s*(\d+)\]\s*(.*)", name_text)
            if m:
                company_id   = m.group(1)
                company_name = m.group(2).strip()
            else:
                company_name = name_text
        except Exception:
            pass

        if show:
            _mat_driver = driver
        else:
            driver.quit()

        result = {
            "serial":          serial,
            "scanner_sn":      scanner_sn,
            "wand_sn":         wand_sn,
            "reg_to":          reg_to,
            "status":          status,
            "scanner_model":   scanner_model,
            "notes":           notes,
            "sw_package":      sw_package,
            "sw_expiry":       sw_expiry,
            "company_id":      company_id,
            "company_name":    company_name,
            "company_address": company_address,
        }
        if not wand_sn:
            result["debug"] = "Wand field empty — check the S/N or the MAT flow"
        return ok(result)

    except Exception as e:
        if driver:
            try: driver.quit()
            except Exception: pass
        return err(f"MAT scrape error: {e}")


def mat_close_browser():
    global _mat_driver
    if _mat_driver is not None:
        try:
            _mat_driver.quit()
        except Exception:
            pass
        _mat_driver = None
    return ok("Browser closed.")


# ─────────────────────────────────────────────
# SALESFORCE (SF) — TICKET CHECKER
# ─────────────────────────────────────────────
# Unlike MAT, this doesn't store credentials — the technician is already
# logged into Salesforce via corporate SSO on this machine. We drive a
# dedicated, always-visible Edge window with its own persistent profile:
# the first use requires a manual SSO login in that window, and every call
# after that reuses the saved session cookies automatically.

_SF_HOME_URL    = "https://aligntech.lightning.force.com/lightning/page/home"
_SF_PROFILE_DIR = os.path.join(os.environ.get("APPDATA", "C:\\Users"), "iTeroToolBox", "sf_profile")
_SF_CFG_PATH    = os.path.join(os.environ.get("APPDATA", "C:\\Users"), "iTeroToolBox", "sf_config.json")
_sf_driver      = None  # kept alive between calls — same window/session reused


def sf_load_settings():
    try:
        if os.path.exists(_SF_CFG_PATH):
            with open(_SF_CFG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
            return ok({"show_browser": bool(cfg.get("show_browser", False))})
    except Exception:
        pass
    return ok({"show_browser": False})


def sf_save_settings(show_browser=False):
    try:
        os.makedirs(os.path.dirname(_SF_CFG_PATH), exist_ok=True)
        with open(_SF_CFG_PATH, "w", encoding="utf-8") as f:
            json.dump({"show_browser": bool(show_browser)}, f)
        return ok(None)
    except Exception as e:
        return err(f"Could not save: {e}")


def _sf_get_driver():
    global _sf_driver
    if _sf_driver is not None:
        try:
            _ = _sf_driver.window_handles  # liveness check — throws if the window was closed
            return _sf_driver
        except Exception:
            _sf_driver = None

    from selenium import webdriver
    from selenium.webdriver.edge.options import Options

    show_browser = False
    try:
        if os.path.exists(_SF_CFG_PATH):
            with open(_SF_CFG_PATH, encoding="utf-8") as f:
                show_browser = bool(json.load(f).get("show_browser", False))
    except Exception:
        pass

    os.makedirs(_SF_PROFILE_DIR, exist_ok=True)
    opts = Options()
    opts.add_argument(f"--user-data-dir={_SF_PROFILE_DIR}")
    opts.add_argument("--profile-directory=Default")
    if not show_browser:
        opts.add_argument("--headless=new")
    # The app runs elevated (admin) — Chromium's sandbox refuses to
    # initialize under Administrator/SYSTEM without --no-sandbox, which
    # crashes the browser instantly ("DevToolsActivePort file doesn't
    # exist"). MAT's launcher already carries these same flags.
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    _sf_driver = webdriver.Edge(options=opts)
    return _sf_driver


def sf_open_home():
    """Open (or focus) the dedicated Salesforce window at the home page —
    used for the first-time manual SSO login, or just to bring it forward."""
    try:
        driver = _sf_get_driver()
    except Exception as e:
        return err(f"Could not open the Salesforce browser: {e}")
    try:
        driver.get(_SF_HOME_URL)
        driver.switch_to.window(driver.current_window_handle)
        return ok(None, msg="Salesforce opened.")
    except Exception as e:
        return err(f"Error opening Salesforce: {e}")


def sf_open_ticket(ticket_number=""):
    """Type a ticket number into Salesforce's global search and press Enter
    to land on the search results page."""
    ticket_number = (ticket_number or "").strip()
    if not ticket_number:
        return err("Enter a ticket number.")

    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.action_chains import ActionChains
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
    except ImportError:
        return err("selenium not installed. Run: pip install selenium")

    try:
        driver = _sf_get_driver()
    except Exception as e:
        return err(f"Could not open the Salesforce browser: {e}")

    try:
        driver.get(_SF_HOME_URL)
        wait = WebDriverWait(driver, 20, ignored_exceptions=[StaleElementReferenceException])

        # Give the SPA a moment to redirect to the SSO/login page if the
        # saved session has expired.
        time.sleep(1.5)
        if "lightning.force.com" not in driver.current_url:
            return ok({"status_text": "login_required"},
                      msg="Log in on the Salesforce window that opened, then try again.")

        # Open the global search box. Salesforce auto-focuses the input the
        # instant it opens — trying to separately locate and click that
        # input ourselves was the bug: with several similar elements in
        # that widget (the "Search: All" scope dropdown sits right next to
        # it), our click was landing on the wrong one and knocking focus
        # off the real field. Instead, just type at whatever the browser
        # currently has focused — no element targeting required.
        search_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "button.search-button, button[aria-label='Search']")))
        search_btn.click()

        typed_value = ""
        for attempt in range(4):
            time.sleep(0.5 + attempt * 0.4)  # give the SPA progressively more settle time
            ActionChains(driver).send_keys(ticket_number).perform()
            typed_value = driver.execute_script(
                "var el = document.activeElement; return el ? (el.value || '') : '';"
            ) or ""
            if ticket_number in typed_value:
                break  # text landed — move on to waiting for suggestions
            # Didn't land (or landed partially) — clear whatever's in the
            # focused field before the next attempt so we don't accumulate
            # garbage across retries.
            ActionChains(driver).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).send_keys(Keys.DELETE).perform()

        if ticket_number not in typed_value:
            return err(f"The search field never received the text after several attempts "
                       f"(ended up with: \"{typed_value}\"). Check the Salesforce window.")

        # Text landed — press Enter right away to run the full search and
        # land on the results page.
        ActionChains(driver).send_keys(Keys.RETURN).perform()

        # The results page renders the matched record's number as a real
        # <a href="/lightning/r/<id>/view" class="outputLookupLink" target=
        # "_blank">. Give it 2.5s to render, then click it. We use a native
        # Selenium click (not a JS-triggered one) so Edge treats it as a
        # trusted user gesture and opens the new tab in the foreground.
        time.sleep(2.5)
        try:
            links = WebDriverWait(driver, 8, ignored_exceptions=[StaleElementReferenceException]).until(
                lambda d: [a for a in d.find_elements(By.CSS_SELECTOR, "a.outputLookupLink[data-recordid]")
                           if a.is_displayed()] or False)
        except TimeoutException:
            return ok({"status_text": "no_case_match"},
                      msg=f"Searched \"{ticket_number}\", but the ticket link never showed up. Check the Salesforce window.")

        link = next((a for a in links
                     if (a.get_attribute("title") or a.text or "").strip() == ticket_number), links[0])

        handles_before = driver.window_handles
        link.click()

        # target="_blank" opens a new tab — follow the driver's focus there
        # so the app (and the technician) end up looking at the ticket. Some
        # orgs also pop a Twilio softphone window around the same time, so
        # among the new handles we specifically pick the one that actually
        # looks like a Salesforce record view rather than just grabbing the
        # last one opened.
        time.sleep(1.2)
        new_handles = [h for h in driver.window_handles if h not in handles_before]
        record_handle = None
        for h in new_handles:
            try:
                driver.switch_to.window(h)
                url = (driver.current_url or "").lower()
                if "lightning.force.com" in url and "twilio" not in url:
                    record_handle = h
                    break
            except Exception:
                continue
        if record_handle is None and new_handles:
            driver.switch_to.window(new_handles[-1])
        elif record_handle is not None:
            driver.switch_to.window(record_handle)

        # Give the record page's LWCs a moment to finish rendering before
        # anything tries to read fields off it.
        time.sleep(3)

        return ok({"status_text": "opened"}, msg=f"Ticket {ticket_number} opened.")

    except Exception as e:
        return err(f"Error searching for the ticket: {e}")


_SF_DEEP_EXTRACT_JS = r"""
function deepQueryAll(sel, root) {
    root = root || document;
    var results = Array.prototype.slice.call(root.querySelectorAll(sel));
    var all = root.querySelectorAll('*');
    for (var i = 0; i < all.length; i++) {
        if (all[i].shadowRoot) {
            results = results.concat(deepQueryAll(sel, all[i].shadowRoot));
        }
    }
    return results;
}

function realParent(node) {
    // node.parentElement is null once you reach the top of a shadow root's
    // content -- the true parent at that point is the shadow HOST, reached
    // via getRootNode().host, not a further .parentElement hop. Without
    // this, climbing stops dead at the first shadow boundary.
    if (node.parentElement) return node.parentElement;
    var root = node.getRootNode();
    if (root && root.host) return root.host;
    return null;
}

function deepText(el) {
    var text = '';
    function walk(node) {
        if (node.nodeType === Node.TEXT_NODE) {
            text += node.textContent;
            return;
        }
        if (node.nodeType !== Node.ELEMENT_NODE) return;
        if (node.shadowRoot) {
            var sc = node.shadowRoot.childNodes;
            for (var i = 0; i < sc.length; i++) walk(sc[i]);
        }
        var kids = node.childNodes;
        for (var j = 0; j < kids.length; j++) walk(kids[j]);
    }
    walk(el);
    return text.trim();
}

function findValueElement(targetName) {
    var containers = deepQueryAll("[data-target-selection-name='" + targetName + "']");
    for (var i = 0; i < containers.length; i++) {
        // Try inside the target container itself first.
        var valEls = deepQueryAll('.test-id__field-value', containers[i]);
        for (var j = 0; j < valEls.length; j++) {
            if (deepText(valEls[j])) return valEls[j];
        }
        var direct = deepQueryAll('[data-output-element-id="output-field"]', containers[i]);
        for (var d = 0; d < direct.length; d++) {
            if (deepText(direct[d])) return direct[d];
        }
        // The actual value is often NOT nested inside this div in the real
        // (unflattened) DOM -- it's a light-DOM child that lives one or more
        // levels up (on the records-record-layout-item host, or further if
        // that host itself sits at a shadow-root boundary), only visually
        // projected into the <slot> placeholder inside. Climb the real
        // parent chain (crossing shadow boundaries) and look there.
        var node = containers[i];
        for (var depth = 0; depth < 3; depth++) {
            node = realParent(node);
            if (!node) break;
            var candidates = deepQueryAll('[slot="outputField"], [data-output-element-id="output-field"]', node);
            for (var k = 0; k < candidates.length; k++) {
                if (deepText(candidates[k])) return candidates[k];
            }
        }
    }
    return null;
}

function fieldValue(targetName) {
    var el = findValueElement(targetName);
    return el ? deepText(el) : '';
}

function addressValue(targetName) {
    var el = findValueElement(targetName);
    if (!el) return '';
    var lines = deepQueryAll('.slds-truncate', el);
    if (lines.length) {
        var parts = lines.map(function(l) { return (l.textContent || '').trim(); }).filter(Boolean);
        if (parts.length) return parts.join(', ');
    }
    return deepText(el);
}

var matContainers = deepQueryAll("[data-target-selection-name='sfdc:RecordField.Account.MAT_ID__c']");
var matValueEls = matContainers.length ? deepQueryAll('.test-id__field-value', matContainers[0]) : [];

// Climb a few real (unflattened) ancestor levels so we can see the ground
// truth of what's actually around this field in the live DOM, instead of
// guessing based on how the DevTools inspector visually flattens slotted
// content.
var matAncestorHtml = '';
if (matContainers.length) {
    var anc = matContainers[0];
    for (var a = 0; a < 3; a++) {
        var up = realParent(anc);
        if (!up) break;
        anc = up;
    }
    matAncestorHtml = anc.outerHTML;
}

return {
    mat_id:  fieldValue("sfdc:RecordField.Account.MAT_ID__c"),
    address: addressValue("sfdc:RecordField.Account.ShippingAddress"),
    asset:   fieldValue("sfdc:RecordField.Case.AssetId"),
    subject: fieldValue("sfdc:RecordField.Case.Subject"),
    debug_container_found: matContainers.length,
    debug_value_el_found: matValueEls.length,
    debug_raw_html: matAncestorHtml.slice(0, 4000)
};
"""


def _sf_search_all_frames(driver, script, max_depth=5):
    """Salesforce Console renders each record panel (Account, Case detail,
    etc.) inside its own nested <iframe> for isolation between Aura/LWC
    contexts. execute_script() only ever sees the top document, so a field
    that's genuinely present in the DOM (confirmed via the browser's own
    inspector) can still be invisible to us if it's one frame down. This
    walks every iframe in the frame tree, re-running the same extraction
    script inside each one, until a frame reports the field found."""
    from selenium.webdriver.common.by import By

    counters = {"frames_visited": 0}

    def search(depth):
        result = driver.execute_script(script) or {}
        if (result.get("mat_id") or "").strip():
            return result
        if depth >= max_depth:
            return result
        try:
            frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
        except Exception:
            frames = []
        for fr in frames:
            try:
                driver.switch_to.frame(fr)
            except Exception:
                continue
            counters["frames_visited"] += 1
            try:
                sub = search(depth + 1)
            except Exception:
                sub = {}
            driver.switch_to.parent_frame()
            if (sub.get("mat_id") or "").strip():
                return sub
        return result

    driver.switch_to.default_content()
    try:
        result = search(0)
        result["_frames_visited"] = counters["frames_visited"]
        return result
    finally:
        driver.switch_to.default_content()


def sf_read_ticket():
    """Read key fields off the currently-open Salesforce ticket (Case) page.

    Being built field-by-field against exact HTML confirmed on the live
    page (data-target-selection-name -> .test-id__field-value), instead of
    guessing selectors. Only MAT ID is wired up so far."""
    global _sf_driver
    if _sf_driver is None:
        return err("No Salesforce window is open. Open a ticket first.")

    driver = _sf_driver
    data = {}
    try:
        for attempt in range(4):
            data = _sf_search_all_frames(driver, _SF_DEEP_EXTRACT_JS) or {}
            if (data.get("mat_id") or "").strip():
                break
            time.sleep(1.0 + attempt * 0.5)
    except Exception as e:
        return err(f"Error reading ticket: {e}")

    mat_id  = (data.get("mat_id") or "").strip()
    address = (data.get("address") or "").strip()
    asset   = (data.get("asset") or "").strip()
    subject = (data.get("subject") or "").strip()
    result = {
        "mat_id":         mat_id or "MAT ID not found.",
        "address":        address or "No address found.",
        "asset":          asset or "No asset found.",
        "asset_source":   "asset" if asset else "",
        "subject":        subject or "No subject found.",
        "order_comments": "No order comments yet.",
    }

    if not mat_id:
        try:
            cur_url = driver.current_url
            cur_title = driver.title
            n_handles = len(driver.window_handles)
        except Exception as e:
            cur_url, cur_title, n_handles = f"<error: {e}>", "", 0

        result["debug_container_found"] = data.get("debug_container_found", 0)
        result["debug_value_el_found"] = data.get("debug_value_el_found", 0)
        result["debug_raw_html"] = data.get("debug_raw_html", "")
        result["debug_frames_visited"] = data.get("_frames_visited", 0)
        result["debug_url"] = cur_url
        result["debug_title"] = cur_title
        result["debug_handles"] = n_handles
        return ok(result, msg="MAT ID not found — see debug box.")

    return ok(result)


def sf_close_browser():
    global _sf_driver
    if _sf_driver is not None:
        try:
            _sf_driver.quit()
        except Exception:
            pass
        _sf_driver = None
    return ok(None, msg="Salesforce window closed.")
